"""Writes the engine's output into the SAME workbook schema (Dashboard,
Positions, Reset Watch are left untouched — those are user-managed)."""
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from . import config as C

FONT = "Arial"
FILLS = {"NEW": "C6EFCE", "PROMOTED": "BDD7EE", "DEMOTED": "FFE699", "SAME": None}
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
FIRST_ROW, LAST_ROW = 4, 40


def _put(ws, r, col, val, color="000000", num=None, align="left", fill=None):
    c = ws.cell(row=r, column=col, value=val)
    c.font = Font(name=FONT, color=color, size=10)
    c.alignment = Alignment(horizontal=align, vertical="center",
                            wrap_text=(col in (5, 24)))
    c.border = BORDER
    if num:
        c.number_format = num
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    return c


def update_workbook(rows_by_tier: dict, trans: dict, path=C.WORKBOOK_PATH):
    wb = load_workbook(path)
    today = date.today()
    for tier in C.TIER_ORDER:
        ws = wb[tier]
        # clear the data region
        for r in range(FIRST_ROW, LAST_ROW + 1):
            for col in range(1, 25):
                ws.cell(row=r, column=col).value = None
                ws.cell(row=r, column=col).fill = PatternFill(fill_type=None)
        rows = rows_by_tier.get(tier, [])
        for i, d in enumerate(rows):
            r = FIRST_ROW + i
            status = trans.get(d["ticker"], "SAME")
            fill = FILLS.get(status)
            _put(ws, r, 1, d["ticker"], fill=fill)
            _put(ws, r, 2, d["name"], fill=fill)
            _put(ws, r, 3, d["exchange"], align="center", fill=fill)
            _put(ws, r, 4, d["sector"], fill=fill)
            _put(ws, r, 5, d["summary"], fill=fill)
            _put(ws, r, 6, d["stage"], align="center", fill=fill)
            _put(ws, r, 7, d["tt"], align="center", fill=fill)
            _put(ws, r, 8, d["rs"], align="center", fill=fill)
            _put(ws, r, 9, d["fund"], align="center", fill=fill)
            _put(ws, r, 10, d["setup"], fill=fill)
            _put(ws, r, 11, d["footprint"], fill=fill)
            _put(ws, r, 12, d["conv"], align="center", fill=fill)
            _put(ws, r, 13, d["pivot_flag"], align="center", fill=fill)
            _put(ws, r, 14, d["buyable_flag"], align="center", fill=fill)
            _put(ws, r, 15, d["pb"], align="center", fill=fill)
            # inputs (blue)
            _put(ws, r, 16, d["pivot"], color="0000FF", num="$#,##0.00", align="right", fill=fill)
            # formulas (black) — re-applied every run
            _put(ws, r, 17, f'=IF(P{r}="","",P{r}*(1+Dashboard!$B$8))', num="$#,##0.00", align="right", fill=fill)
            _put(ws, r, 18, d["entry"], color="0000FF", num="$#,##0.00", align="right", fill=fill)
            _put(ws, r, 19, d["stop"], color="0000FF", num="$#,##0.00", align="right", fill=fill)
            _put(ws, r, 20, f'=IF(OR(R{r}="",S{r}="",R{r}=0),"",(R{r}-S{r})/R{r})', num="0.0%", align="right", fill=fill)
            _put(ws, r, 21, f'=IF(OR(R{r}="",S{r}="",R{r}=S{r}),"",ROUND(Dashboard!$B$7/(R{r}-S{r}),0))', num="#,##0", align="right", fill=fill)
            _put(ws, r, 22, d.get("added", today), num="yyyy-mm-dd", align="center", fill=fill)
            _put(ws, r, 23, today, num="yyyy-mm-dd", align="center", fill=fill)
            _put(ws, r, 24, f'[{status}] ' + d["reason"], fill=fill)
    wb.save(path)
